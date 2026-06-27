"""
Script de migration : importe les données de l'ancien fichier Excel
(PROJET_PRIVE_PCA.xlsx) vers la base de données du logiciel.

Utilisation :
    python import_excel.py "chemin/vers/PROJET_PRIVE_PCA.xlsx"

Logique :
- Les onglets "TABLEAU DE CHEQUE...", "TABLEAU n", "TC n" contiennent des lignes
  N° / A L'ORDRE DE / OBJETS / MONTANT.
- Si l'objet de la ligne contient un mot-clé d'avance (avance, solde, avenant,
  acompte), la ligne est enregistrée comme une AVANCE pour le prestataire
  correspondant (créé automatiquement s'il n'existe pas encore).
- Sinon, la ligne est enregistrée comme un ACHAT (fournisseur = "A L'ORDRE DE").
- La date du tableau (cellule D1 de chaque onglet) est utilisée comme date.
"""
import sys
import re
import sqlite3
import openpyxl
from datetime import datetime
import os

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "data", "gestion_chantiers.db")

AVANCE_KEYWORDS = ["avance", "solde", "avenant", "acompte", "régulation", "regularisation"]

SKIP_SHEETS = {"RECAP MO", "SITUATION", "Feuil1"}


def is_avance(objet):
    if not objet:
        return False
    o = objet.lower()
    return any(k in o for k in AVANCE_KEYWORDS)


def clean_amount(v):
    try:
        if v is None or str(v).strip() == "":
            return 0.0
        if isinstance(v, str) and v.strip().startswith("#"):
            return 0.0
        return float(v)
    except (ValueError, TypeError):
        return 0.0


def get_or_create_prestataire(conn, nom, default_chantier_id):
    nom = nom.strip()
    row = conn.execute("SELECT id FROM prestataires WHERE nom = ? COLLATE NOCASE", (nom,)).fetchone()
    if row:
        return row[0]
    cur = conn.execute("""
        INSERT INTO prestataires (chantier_id, nom, montant_contrat, statut, date_creation)
        VALUES (?,?,?,?,?)
    """, (default_chantier_id, nom, 0, "En cours", datetime.now().isoformat()))
    return cur.lastrowid


def main(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA foreign_keys=ON")

    chantier = conn.execute("SELECT id FROM chantiers LIMIT 1").fetchone()
    if not chantier:
        cur = conn.execute("INSERT INTO chantiers (nom, description, date_creation) VALUES (?,?,?)",
                            ("Chantier importé", "Créé automatiquement par l'import", datetime.now().isoformat()))
        chantier_id = cur.lastrowid
    else:
        chantier_id = chantier[0]

    nb_achats, nb_avances, nb_prestataires_crees = 0, 0, 0
    noms_avant = {r[0] for r in conn.execute("SELECT nom FROM prestataires").fetchall()}

    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue
        ws = wb[sheet_name]

        # date du tableau (D1) si présente
        date_tableau = None
        d1 = ws.cell(row=1, column=4).value
        if isinstance(d1, datetime):
            date_tableau = d1.strftime("%Y-%m-%d")

        last_nom = None
        for row in ws.iter_rows(min_row=3, values_only=False):
            cells = [c.value for c in row[:4]]
            if len(cells) < 4:
                continue
            num, nom, objet, montant = cells[0], cells[1], cells[2], cells[3]

            if nom and isinstance(nom, str) and nom.strip():
                last_nom = nom.strip()
            elif not nom:
                nom = last_nom

            if not objet or not nom:
                continue
            if str(num).strip().upper().startswith(("MONTANT TOTAL", "TOTAL")):
                continue

            montant_f = clean_amount(montant)
            objet_str = str(objet).strip()

            if is_avance(objet_str):
                pid = get_or_create_prestataire(conn, nom, chantier_id)
                conn.execute("""
                    INSERT INTO avances (prestataire_id, date, montant, libelle, mode_paiement, saisi_par, date_saisie)
                    VALUES (?,?,?,?,?,?,?)
                """, (pid, date_tableau, montant_f, objet_str, "Chèque", "import_excel", datetime.now().isoformat()))
                nb_avances += 1
            else:
                if montant_f == 0:
                    continue
                conn.execute("""
                    INSERT INTO achats (chantier_id, date, fournisseur, objet, categorie, montant,
                        tableau_ref, mode_paiement, saisi_par, date_saisie)
                    VALUES (?,?,?,?,?,?,?,?,?,?)
                """, (chantier_id, date_tableau, nom, objet_str, None, montant_f,
                      sheet_name, "Chèque", "import_excel", datetime.now().isoformat()))
                nb_achats += 1

    conn.commit()
    noms_apres = {r[0] for r in conn.execute("SELECT nom FROM prestataires").fetchall()}
    nb_prestataires_crees = len(noms_apres - noms_avant)
    conn.close()

    print("="*60)
    print(" IMPORT TERMINÉ")
    print(f" Achats importés        : {nb_achats}")
    print(f" Avances importées       : {nb_avances}")
    print(f" Prestataires créés      : {nb_prestataires_crees}")
    print("="*60)
    print(" ⚠ Pensez à ouvrir chaque prestataire importé pour renseigner :")
    print("   - le montant du contrat")
    print("   - le téléphone")
    print("   - les dates de début / fin")
    print("   (ces informations n'existaient pas dans l'ancien fichier Excel)")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage : python import_excel.py chemin/vers/votre_fichier.xlsx")
        sys.exit(1)
    if not os.path.exists(os.path.join(BASE_DIR, "data")):
        os.makedirs(os.path.join(BASE_DIR, "data"))
    if not os.path.exists(DB_PATH):
        # initialiser la base si elle n'existe pas encore
        sys.path.insert(0, BASE_DIR)
        from app import init_db
        init_db()
    main(sys.argv[1])
